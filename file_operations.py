import openpyxl

# Function to load attendance data from the file
def load_attendance_from_file(filename="attendance.xlsx"):
    try:
        # Load the workbook and access the active sheet
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        # Read all rows from the sheet (skipping header)
        attendance_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            student, attendance = row
            attendance_data.append({"student": student, "attendance": attendance})

        return attendance_data
    except FileNotFoundError:
        print(f"File '{filename}' not found.")
        return []
    except Exception as e:
        print(f"Error loading file: {e}")
        return []

# Function to save attendance data to the file
def save_attendance_to_file(attendance_data, filename="attendance.xlsx"):
    try:
        # Load the workbook or create a new one if not exists
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Student", "Attendance"])  # Add header if creating a new file

        # Clear existing data (if necessary) and write the new data
        sheet.delete_rows(2, sheet.max_row)  # Remove old attendance data
        for record in attendance_data:
            sheet.append([record["student"], record["attendance"]])

        # Save the workbook
        workbook.save(filename)
        print(f"Attendance saved to '{filename}' successfully.")
    except Exception as e:
        print(f"Error saving file: {e}")